"open the serum uv data"

from typing import BinaryIO, Union, Callable, Dict, Optional
from enum import Enum
from abc import ABC, abstractmethod
import re


from openpyxl import load_workbook  # type: ignore
from openpyxl.workbook import Workbook
from openpyxl.worksheet.worksheet import Worksheet

from boxy.models import ParsedData, Plate, Well  # noqa
from boxy.base_parser import BaseParser

import pandas as pd


class SynergyApplication(ABC):
    """
    The base class for all Lunatic applications.
    """

    def __init__(self, name: str, version: str | None = None):
        self.name = name
        self.version = version

    @abstractmethod
    def header_conversion_dictionary(self, header: list[str]) -> dict[str, str]:
        """
        Converts the header of a Lunatic application to a dictionary of the column names and their corresponding values.
        Args:
            header (list[str]): The header of the Lunatic application.
        Returns:
            dict[str, str]: A dictionary of the column names and their corresponding values.
        """
        pass


class KineticTimeseries(SynergyApplication):
    """
    The class for the raw absorbance application of the Lunatic instrument.
    """

    def __init__(self, version=None):
        super().__init__("Raw Absorbance", version)

    def header_conversion_dictionary(self, header: list[str]) -> dict[str, str]:
        """
        Provides a dictionary for changing the column names of the raw absorbance application to just wavelength values.
        Args:
            header (list[str]): The header of the raw absorbance application.
        Returns:
            dict[str, str]: A dictionary of the column names and their corresponding values.
        """
        return dict(
            (column, column.split()[0][1:])
            for column in header
            if re.match(r"A\d+ \(10mm\)", column)
        )


class SynergyApplications(Enum):
    """
    Enum class for the different applications of the Lunatic instrument.
    Attributes
    ----------
    RAW_ABSORBANCE : str
        The raw absorbance application.
    """

    RAW_ABSORBANCE = KineticTimeseries


class ExcelParser(BaseParser):
    "All the Excel stuff"
    _well_column: str | None = None
    _well_string_to_row: Callable[[str], int] = lambda x: ord(x[0]) - 65
    _well_string_to_col: Callable[[str], int] = lambda x: int(x[1:]) - 1

    def __init_subclass__(cls, **kwargs):
        super().__init_subclass__(**kwargs)

        if cls._well_column is None:
            raise TypeError(
                f"Class variable 'well_column' must be set in {cls.__name__}"
            )

        if not isinstance(cls._well_string_to_row, staticmethod):
            cls._well_string_to_row = staticmethod(cls._well_string_to_row)

        if not isinstance(cls._well_string_to_col, staticmethod):
            cls._well_string_to_col = staticmethod(cls._well_string_to_col)

    @staticmethod
    def find_row_starting_with(sheet, value):
        """
        Finds the first row in the given sheet that starts with the specified value.
        Args:
            sheet (openpyxl.worksheet.worksheet.Worksheet): The worksheet to search through.
            value (str): The value to search for in the first cell of each row.
        Returns:
            int: The row number of the first row that starts with the specified value.
        Raises:
            StopIteration: If no row starts with the specified value.
        """

        return next(row[0].row for row in sheet.iter_rows() if row[0].value == value)

    @staticmethod
    def find_next_empty_row(sheet, start_row: int = 1):
        """
        Finds the next empty row in an Excel sheet starting from a specified row.

        Args:
            sheet: The Excel sheet object to search through.
            start_row (int, optional): The row number to start searching from. Defaults to 1.

        Returns:
            int: The row number of the next empty row.

        Raises:
            StopIteration: If no empty row is found.
        """
        return next(
            row[0].row
            for row in sheet.iter_rows(min_row=start_row)
            if all(cell.value is None for cell in row)
        )

    @staticmethod
    def _metadata_from_cell_range(sheet, start_row, end_row, start_col, end_col):
        """
        Extracts metadata from a range of cells in an Excel sheet.

        Args:
            sheet: The Excel sheet object to extract metadata from.
            start_row (int): The starting row number of the cell range.
            end_row (int): The ending row number of the cell range.
            start_col (int): The starting column number of the cell range.
            end_col (int): The ending column number of the cell range.

        Returns:
            dict: A dictionary containing the extracted metadata.
        """
        metadata = {}
        for row in sheet.iter_rows(
            min_row=start_row, max_row=end_row, min_col=start_col, max_col=end_col
        ):
            for cell in row:
                if cell.value is not None:
                    metadata[cell.value] = cell.offset(column=1).value
                    # advance to the next row
                    break
        return metadata

    def _well_df_from_cell_range(
        self,
        sheet,
        start_row,
        end_row,
        start_col,
        headers: list[str] | None = None,
    ):
        """
        Extracts well data from a range of cells in an Excel sheet.

        Args:
            sheet: The Excel sheet object to extract well data from.
            start_row (int): The starting row number of the cell range.
            end_row (int): The ending row number of the cell range.
            start_col (int): The starting column number of the cell range.
            end_col (int): The ending column number of the cell range.

        Returns:
            pd.DataFrame: A DataFrame containing the extracted well data.
        """
        if headers is None:
            headers = [cell.value for cell in sheet[start_row]]
        well_df = pd.DataFrame(
            sheet.iter_rows(
                min_row=start_row + 1,
                max_row=end_row,
                min_col=start_col,
                max_col=start_col + len(headers),
                values_only=True,
            ),
            columns=headers,
        )
        if "row" not in well_df.columns:
            well_df["row"] = None
        well_df.loc[:, "row"] = well_df[self._well_column].apply(
            self._well_string_to_row
        )
        if "col" not in well_df.columns:
            well_df["col"] = None
        well_df.loc[:, "col"] = well_df[self._well_column].apply(
            self._well_string_to_col
        )
        return well_df

    @staticmethod
    def _get_xlsx_workbook(file: Union[str, BinaryIO]) -> Workbook:
        """
        Opens and Excel xlsx file and returns the workbook for the spreadsheet
        :param file: file path string, or bytes-like file object
        :return workbook: openpyxl Workbook object
        """

        workbook = load_workbook(file)
        return workbook

    @staticmethod
    def _get_first_xlsx_worksheet(
        file: Union[str, BinaryIO], sheet_name: Optional[str] = None
    ) -> Worksheet:
        """
        Opens an Excel xlsx file and returns the first worksheet
        :param file: file path string, or bytes-like file object
        :param sheet_name: name of sheet to return
        :return sheet: openpyxl Worksheet object for first sheet in xlsx
        """

        workbook = load_workbook(file)
        if sheet_name is None:
            sheet = workbook.active
        else:
            if sheet_name in workbook.sheetnames:
                sheet = workbook[sheet_name]
            else:
                raise ValueError(f"Cannot find {sheet_name} in the file")
        return sheet


def _is_well(name: str) -> bool:
    """
    Check if a name is a well name.
    Args:
        name (str): The name to check.

    Returns:
        bool: True if the name is a well name, False otherwise.
    """
    if not isinstance(name, str):
        return False
    if re.match(r"[A-Z]{1,2}\d{1,2}", name) is not None:
        return True
    return False


def _has_wells(names: list, cutoff: float = 0.5) -> bool:
    """
    Check if a list of names contains well names.
    Args:
        names (list): A list of names to check.
        cutoff (float, optional): The minimum proportion of well names required. Defaults to 0.5.

    Returns:
        bool: True if the list contains well names, False otherwise.
    """
    wells = [name for name in names if _is_well(name)]
    return (len(wells) / len(names)) > cutoff


def _create_well_table(
    table: pd.DataFrame, index_name: str | None = None
) -> pd.DataFrame:
    """
    Create a well table from a DataFrame. A well table has wells as columns and time or wavelength as the index.
    Args:
        table (pd.DataFrame): The input DataFrame.
    Returns:
        pd.DataFrame: The well table with wells as columns and time or wavelength as the index.
    """
    if not isinstance(table, pd.DataFrame):
        raise TypeError("table must be a pandas DataFrame")
    if table.empty:
        raise ValueError("table is empty")
    # check if the first row contains well names
    if _has_wells(table.iloc[0], 0.5):
        table = pd.DataFrame(data=table.values[1:, :], columns=table.values[0, :])
    elif _has_wells(table.iloc[:, 0], 0.5):
        table = pd.DataFrame(data=table.values[:, 1:].T, columns=table.values[:, 0])
    else:
        raise ValueError("Table does not contain well names")
    if index_name is not None:
        table.index = table[index_name]
        table = table.drop(columns=[index_name])
    else:
        index_priorities = ["time", "wavelength"]
        for index_priority in index_priorities:
            if index_priority in [
                x.lower() for x in table.columns if isinstance(x, str)
            ]:
                index_column = [
                    x for x in table.columns if x.lower() == index_priority
                ][0]
                table.index = table[index_column]
                table = table.drop(columns=[index_column])
                break
    return table


class SynergyParser(ExcelParser):
    """
    LunaticParser is a parser class that extends the BaseParser to parse data from an Excel file.
    Methods
    -------
    parse(file: Union[str, BinaryIO]) -> ParsedData
        Parses the given Excel file and extracts metadata and sample data.
    parse(file: Union[str, BinaryIO]) -> ParsedData
        Parameters
        ----------
        file : Union[str, BinaryIO]
            The path to the Excel file or a file-like object containing the Excel data.
        Returns
        -------
        ParsedData
            An object containing the extracted metadata and sample data.
        Raises
        ------
        ValueError
            If there is an issue with reading the Excel file or finding the required data.
    """

    _well_column = "Plate Position"

    def parse(self, file: Union[str, BinaryIO]) -> ParsedData:
        """
        Parses the given Excel file and extracts metadata and sample data.
        Parameters
        ----------
        file : Union[str, BinaryIO]

            The path to the Excel file or a file-like object containing the Excel data.
        Returns
        -------
        ParsedData

            An object containing the extracted metadata and sample data.
        Raises
        ------
        ValueError
            If there is an issue with reading the Excel file or finding the required data.
        """
        # file = "uv/data/UV serum.xlsx"
        try:
            sheet = self._get_first_xlsx_worksheet(file)
        except ValueError as e:
            raise ValueError(e) from e
        # split the sheet into tables separated by empty rows
        section_starts = [
            (row[0].row + 3, row[0].offset(row=1).value)
            for row in sheet.iter_rows()
            if all(cell.value is None for cell in row)
            and all(cell.offset(row=2).value is None for cell in row)
            and row[0].offset(row=1).value is not None
        ]
        metadata = self._metadata_from_cell_range(
            sheet, 1, section_starts[1][0] - 2, 0, 2
        )
        tables = {
            str(table_name): pd.DataFrame(
                sheet.iter_rows(
                    start_row,
                    (
                        (section_starts[i + 1][0] - 4)
                        if i + 1 < len(section_starts)
                        else sheet.max_row
                    ),
                    2,
                    values_only=True,
                )
            )
            .dropna(how="all")
            .dropna(how="all", axis=1)
            for i, (start_row, table_name) in enumerate(section_starts)
            if table_name != "Software Version"
        }

        well_tables: Dict[str, pd.DataFrame] = {}
        for table_name, table in tables.items():
            try:
                well_tables[table_name] = _create_well_table(table)
            except ValueError:
                pass
            except TypeError:
                pass

        long_data = pd.concat(
            [
                pd.DataFrame(
                    data=table.stack().reset_index().values,
                    columns=["time", "well", "value"],
                ).assign(measurement=table_name)
                for table_name, table in well_tables.items()
            ]
        )

        wells = [
            Well(
                well_location=well,
                row=self._well_string_to_row(well),
                column=self._well_string_to_col(well),
                sample=None,
                annotations=well_data.pivot(
                    index="time", columns="measurement", values="value"
                ).to_dict(orient="list"),
            )
            for well, well_data in long_data.groupby("well")
            if _is_well(well)
        ]

        # Synergy only ever puts one plate per sheet. This is not currently designed for multi-plate parsing.
        # Toward multiplate parsing, we would need to decide how to represent metadata that changes across plates
        plates = [Plate(id=metadata["Plate Number"], wells=wells)]

        return ParsedData(
            metadata=metadata,
            plates=plates,
        )
